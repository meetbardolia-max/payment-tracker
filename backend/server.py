from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import re
import secrets
from io import BytesIO
from datetime import datetime, timezone, timedelta, date
from typing import Optional, List

import bcrypt
import jwt
import openpyxl
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel

UPLOAD_DIR = Path("/app/uploads/cheques")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Sripati Processors Collection Desk")
api_router = APIRouter(prefix="/api")
JWT_ALGORITHM = "HS256"

DELAY_REASONS = [
    "Payment approval pending",
    "Owner / decision-maker unavailable",
    "Person out of station",
    "Cash-flow issue",
    "Payment processing delay",
    "Bank issue",
    "Invoice query / dispute",
    "Quality issue",
    "Quantity issue",
    "Rate issue",
    "Cheque will be issued later",
    "Payment already processed",
    "Other",
]


def now():
    return datetime.now(timezone.utc)


def money(v):
    try:
        return round(float(v or 0), 2)
    except Exception:
        return 0.0


def clean(doc):
    if not doc:
        return None
    doc = dict(doc)
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    for k, v in list(doc.items()):
        if isinstance(v, (datetime, date)):
            doc[k] = v.isoformat()
    return doc


def hash_password(v):
    return bcrypt.hashpw(v.encode(), bcrypt.gensalt()).decode()


def verify_password(v, h):
    return bcrypt.checkpw(v.encode(), h.encode())


def make_token(u):
    return jwt.encode(
        {"sub": u["id"], "email": u["email"], "role": u["role"], "exp": now() + timedelta(hours=8)},
        os.environ["JWT_SECRET"],
        algorithm=JWT_ALGORITHM,
    )


class LoginInput(BaseModel):
    email: str
    password: str


class FollowUpInput(BaseModel):
    snapshot_id: str
    party_id: str
    outcome: str  # paid_full | paid_partial | not_paid
    amount_received: Optional[float] = None
    remaining_amount: Optional[float] = None
    reason: Optional[str] = ""
    reason_other: Optional[str] = ""
    notes: Optional[str] = ""
    next_followup_date: Optional[str] = None
    payment_method: Optional[str] = None
    cheque_number: Optional[str] = None
    cheque_date: Optional[str] = None
    cheque_image: Optional[str] = None
    promise_date: Optional[str] = None
    promise_amount: Optional[float] = None


class AssignmentInput(BaseModel):
    snapshot_id: str
    party_ids: List[str]
    officer_id: str


class CommitInput(BaseModel):
    parties: List[dict]
    period_label: Optional[str] = None
    report_period: Optional[str] = ""
    source_file: str = "outstanding.xlsx"
    total_outstanding: float = 0
    total_bill_amt: float = 0
    total_received: float = 0


async def current_user(request: Request):
    raw = request.cookies.get("access_token") or request.headers.get("Authorization", "").replace("Bearer ", "")
    if not raw:
        raise HTTPException(401, "Please sign in to continue")
    try:
        payload = jwt.decode(raw, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
    except Exception:
        raise HTTPException(401, "Your session has expired")
    u = await db.users.find_one({"id": payload.get("sub"), "active": True}, {"_id": 0})
    if not u:
        raise HTTPException(401, "User not found")
    return clean(u)


def role_guard(*roles):
    async def guard(user=Depends(current_user)):
        if user["role"] not in roles:
            raise HTTPException(403, "You do not have permission for this action")
        return user
    return guard


# ================== EXCEL PARSER ==================
def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0.0
        try:
            return float(s.replace(",", ""))
        except Exception:
            return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _s(v):
    return "" if v is None else str(v).strip()


def parse_outstanding(file_bytes: bytes, filename: str):
    """Parse the Masterwise Groupwise Partywise Detailed Outstanding Report."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    parties = []
    master = None
    group = None
    party_meta = None
    bills: list = []
    party_tot: dict = {}
    report_period = ""

    def flush():
        nonlocal party_meta, bills, party_tot
        if party_meta and bills:
            # Prefer authoritative Pty Tot values from the report itself; fall back to summing bills.
            authoritative = bool(party_tot)
            total_bill = party_tot.get("bill_amt") if authoritative else sum(b["bill_amt"] for b in bills)
            total_rcvd = party_tot.get("rcvd_amt") if authoritative else sum(b["rcvd_amt"] for b in bills)
            total_os = party_tot.get("bill_os") if authoritative else sum(b["bill_os"] for b in bills)
            total_taka = party_tot.get("taka") if authoritative else sum(b["taka"] for b in bills)
            total_mtrs = party_tot.get("mtrs") if authoritative else sum(b["mtrs"] for b in bills)
            parties.append({
                "id": secrets.token_hex(8),
                "master": master or "—",
                "group": party_meta.get("group") or group or "—",
                "party_name": party_meta["name"],
                "party_code": party_meta.get("code"),
                "mobile": party_meta.get("mobile"),
                "address": party_meta.get("address"),
                "bills": list(bills),
                "total_outstanding": round(total_os or 0, 2),
                "total_bill_amt": round(total_bill or 0, 2),
                "total_received": round(total_rcvd or 0, 2),
                "total_taka": round(total_taka or 0, 2),
                "total_mtrs": round(total_mtrs or 0, 2),
                "bill_count": len(bills),
                "authoritative": authoritative,
            })
        party_meta = None
        bills = []
        party_tot = {}

    for r in range(1, ws.max_row + 1):
        a = _s(ws.cell(r, 1).value)
        b = _s(ws.cell(r, 2).value)

        if a.startswith("Reporting"):
            report_period = a
            continue
        if a.startswith("# Master :"):
            flush()
            master = a.split(":", 1)[1].strip()
            continue
        if a.startswith("## Group :"):
            grp = a.split(":", 1)[1].strip()
            group = grp.split(",")[0].strip()
            continue
        if a.startswith("### Party :"):
            flush()
            pty = a.split(":", 1)[1].strip()
            m = re.match(r"(.+?)\s*\(([^)]+)\)\s*,?\s*Mobile\s*:\s*(.+)", pty)
            if m:
                party_meta = {"name": m.group(1).strip(), "code": m.group(2).strip(),
                              "mobile": m.group(3).strip(), "group": group}
            else:
                party_meta = {"name": pty, "code": None, "mobile": None, "group": group}
            continue
        if a.startswith("Address:"):
            if party_meta:
                party_meta["address"] = a[len("Address:"):].strip()
            continue

        # Authoritative party total row from the report itself
        if b == "Pty Tot:" and party_meta is not None:
            party_tot = {
                "taka": _num(ws.cell(r, 8).value),
                "mtrs": _num(ws.cell(r, 9).value),
                "amount": _num(ws.cell(r, 11).value),
                "bill_amt": _num(ws.cell(r, 12).value),
                "rcvd_amt": _num(ws.cell(r, 20).value),
                "bill_os": _num(ws.cell(r, 21).value),
            }
            continue

        # detail bill row: col B has DD/MM/YYYY
        if re.match(r"^\d{2}/\d{2}/\d{4}$", b):
            try:
                d = datetime.strptime(b, "%d/%m/%Y").date()
            except Exception:
                continue
            bucket = "1-15" if d.day <= 15 else "16-31"
            bill_amt = _num(ws.cell(r, 12).value)
            rcvd_amt = _num(ws.cell(r, 20).value)
            raw_os = _num(ws.cell(r, 21).value)
            # For display, clamp per-row outstanding to bill_amt - received so running-balance
            # artifacts (negatives, or values > bill_amt) don't confuse users.
            clean_os = max(0.0, min(raw_os, bill_amt - rcvd_amt)) if bill_amt else raw_os
            bills.append({
                "date": d.isoformat(),
                "display_date": b,
                "period_bucket": bucket,
                "period_label": f"{d.strftime('%b %Y')} · {bucket}",
                "type": _s(ws.cell(r, 3).value) or "PC",
                "taka": _num(ws.cell(r, 8).value),
                "mtrs": _num(ws.cell(r, 9).value),
                "rate": _num(ws.cell(r, 10).value),
                "amount": _num(ws.cell(r, 11).value),
                "bill_amt": bill_amt,
                "tax_cgst": _num(ws.cell(r, 13).value),
                "tax_sgst": _num(ws.cell(r, 14).value),
                "rcvd_amt": rcvd_amt,
                "bill_os": clean_os,
                "raw_bill_os": raw_os,
                "remark": _s(ws.cell(r, 29).value),
            })
            continue
        # ignore Mth Tot / Grp Tot / Mst Tot rows
    flush()

    return {
        "parties": parties,
        "report_period": report_period,
        "source_file": filename,
        "party_count": len(parties),
        "total_outstanding": round(sum(p["total_outstanding"] for p in parties), 2),
        "total_bill_amt": round(sum(p["total_bill_amt"] for p in parties), 2),
        "total_received": round(sum(p["total_received"] for p in parties), 2),
    }


# ================== ROUTES ==================
@api_router.get("/")
async def root():
    return {"message": "Sripati Collection Desk API", "status": "ready"}


@api_router.post("/auth/login")
async def login(data: LoginInput, response: Response):
    email = data.email.strip().lower()
    u = await db.users.find_one({"email": email}, {"_id": 0})
    if not u or not verify_password(data.password, u["password_hash"]):
        raise HTTPException(401, "Incorrect email or password")
    u = clean(u)
    response.set_cookie("access_token", make_token(u), httponly=True, secure=True, samesite="none",
                        max_age=28800, path="/")
    await db.audit_logs.insert_one({"id": secrets.token_hex(8), "action": "login",
                                    "user_id": u["id"], "user_name": u["name"], "created_at": now()})
    return u


@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api_router.get("/auth/me")
async def me(user=Depends(current_user)):
    return user


@api_router.get("/reasons")
async def reasons(user=Depends(current_user)):
    return DELAY_REASONS


@api_router.get("/officers")
async def officers(user=Depends(current_user)):
    docs = await db.users.find({"role": "field_officer", "active": True}, {"_id": 0, "password_hash": 0}).to_list(50)
    return [clean(d) for d in docs]


@api_router.post("/outstanding/upload")
async def upload_preview(file: UploadFile = File(...), user=Depends(role_guard("owner", "head_officer"))):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an .xlsx outstanding report")
    content = await file.read()
    if len(content) > 12 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 12MB)")
    try:
        parsed = parse_outstanding(content, file.filename)
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")
    dup = await db.snapshots.find_one({"source_file": file.filename}, {"_id": 0, "id": 1})
    parsed["duplicate_warning"] = bool(dup)
    return parsed


@api_router.post("/outstanding/commit")
async def upload_commit(payload: CommitInput, user=Depends(role_guard("owner", "head_officer"))):
    if not payload.parties:
        raise HTTPException(400, "No parties to commit")
    snapshot_id = secrets.token_hex(10)
    period_label = payload.period_label or now().strftime("As of %d %b %Y")
    await db.snapshots.update_many({"active": True}, {"$set": {"active": False}})
    snap = {
        "id": snapshot_id,
        "period_label": period_label,
        "source_file": payload.source_file,
        "uploaded_by": user["id"],
        "uploaded_by_name": user["name"],
        "uploaded_at": now().isoformat(),
        "active": True,
        "party_count": len(payload.parties),
        "total_outstanding": payload.total_outstanding,
        "total_bill_amt": payload.total_bill_amt,
        "total_received": payload.total_received,
        "report_period": payload.report_period,
    }
    await db.snapshots.insert_one(dict(snap))
    for p in payload.parties:
        p = dict(p)
        p["snapshot_id"] = snapshot_id
        p.setdefault("assigned_officer_id", None)
        p.setdefault("assigned_officer_name", None)
        await db.snapshot_parties.insert_one(p)
    await db.audit_logs.insert_one({
        "id": secrets.token_hex(8), "action": "snapshot_committed",
        "user_id": user["id"], "user_name": user["name"],
        "snapshot_id": snapshot_id, "party_count": len(payload.parties),
        "created_at": now(),
    })
    return clean(snap)


@api_router.get("/snapshots")
async def snapshots(user=Depends(current_user)):
    docs = await db.snapshots.find({}, {"_id": 0}).sort("uploaded_at", -1).to_list(100)
    return [clean(d) for d in docs]


@api_router.get("/snapshots/active")
async def active_snapshot(user=Depends(current_user)):
    s = await db.snapshots.find_one({"active": True}, {"_id": 0})
    return clean(s) if s else None


@api_router.get("/snapshots/{snapshot_id}/parties")
async def snapshot_parties(snapshot_id: str, search: str = "", view: str = "party",
                           sort: str = "outstanding_desc", user=Depends(current_user)):
    parties = await db.snapshot_parties.find({"snapshot_id": snapshot_id}, {"_id": 0}).to_list(3000)
    if user["role"] == "field_officer":
        parties = [p for p in parties if p.get("assigned_officer_id") == user["id"]]
    q = search.lower().strip()
    if q:
        parties = [p for p in parties if
                   q in (p.get("party_name", "") or "").lower()
                   or q in (p.get("party_code", "") or "").lower()
                   or q in (p.get("master", "") or "").lower()
                   or q in (p.get("group", "") or "").lower()]

    def code_str(p): return (p.get("party_code") or "").upper()
    def name_str(p): return (p.get("party_name") or "").upper()

    def sort_parties(items):
        if sort == "outstanding_desc":
            return sorted(items, key=lambda p: p.get("total_outstanding") or 0, reverse=True)
        if sort == "outstanding_asc":
            return sorted(items, key=lambda p: p.get("total_outstanding") or 0)
        if sort in ("code_asc", "code_desc"):
            with_code = [p for p in items if code_str(p)]
            without = [p for p in items if not code_str(p)]
            with_code.sort(key=code_str, reverse=(sort == "code_desc"))
            return with_code + without
        if sort in ("name_asc", "name_desc"):
            return sorted(items, key=name_str, reverse=(sort == "name_desc"))
        return sorted(items, key=lambda p: p.get("total_outstanding") or 0, reverse=True)

    if view == "master":
        groups: dict = {}
        for p in parties:
            key = p.get("master") or "—"
            g = groups.setdefault(key, {
                "master": key, "parties": [], "total_outstanding": 0,
                "total_bill_amt": 0, "total_received": 0, "party_count": 0,
            })
            g["parties"].append({
                "id": p["id"], "party_name": p["party_name"],
                "party_code": p.get("party_code"), "mobile": p.get("mobile"),
                "total_outstanding": p["total_outstanding"], "bill_count": p["bill_count"],
                "assigned_officer_name": p.get("assigned_officer_name"),
            })
            g["total_outstanding"] += p["total_outstanding"]
            g["total_bill_amt"] += p["total_bill_amt"]
            g["total_received"] += p["total_received"]
            g["party_count"] += 1
        for g in groups.values():
            g["parties"] = sort_parties(g["parties"])
            g["total_outstanding"] = round(g["total_outstanding"], 2)
            g["total_bill_amt"] = round(g["total_bill_amt"], 2)
            g["total_received"] = round(g["total_received"], 2)
        return sorted(groups.values(), key=lambda x: -x["total_outstanding"])

    out = []
    for p in parties:
        out.append({k: v for k, v in p.items() if k != "bills"})
    return sort_parties(out)


@api_router.get("/parties/{party_id}")
async def party_detail(party_id: str, user=Depends(current_user)):
    p = await db.snapshot_parties.find_one({"id": party_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Party not found")
    if user["role"] == "field_officer" and p.get("assigned_officer_id") != user["id"]:
        raise HTTPException(403, "This party is not assigned to you")
    fu = await db.follow_ups.find({"party_id": party_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    p["follow_ups"] = [clean(x) for x in fu]
    return clean(p)


@api_router.post("/follow-ups")
async def create_follow_up(data: FollowUpInput, user=Depends(current_user)):
    p = await db.snapshot_parties.find_one({"id": data.party_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Party not found")
    if user["role"] == "field_officer" and p.get("assigned_officer_id") != user["id"]:
        raise HTTPException(403, "This party is not assigned to you")
    fu = data.model_dump()
    fu["id"] = secrets.token_hex(10)
    fu["party_name"] = p["party_name"]
    fu["party_code"] = p.get("party_code")
    fu["master"] = p.get("master")
    fu["officer_id"] = user["id"]
    fu["officer_name"] = user["name"]
    fu["created_at"] = now().isoformat()
    fu["promise_status"] = "Open" if data.promise_date else None
    if data.reason == "Other" and data.reason_other:
        fu["reason"] = data.reason_other
    await db.follow_ups.insert_one(dict(fu))
    await db.audit_logs.insert_one({
        "id": secrets.token_hex(8), "action": "follow_up_created",
        "user_id": user["id"], "user_name": user["name"],
        "party_id": data.party_id, "party_name": p["party_name"],
        "outcome": data.outcome, "created_at": now(),
    })
    return clean(fu)


@api_router.post("/follow-ups/cheque")
async def upload_cheque(file: UploadFile = File(...), user=Depends(current_user)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "Please upload an image (jpg/png)")
    ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        ext = ".jpg"
    fname = f"{secrets.token_hex(8)}{ext}"
    path = UPLOAD_DIR / fname
    content = await file.read()
    if len(content) > 6 * 1024 * 1024:
        raise HTTPException(400, "Image too large (max 6MB)")
    path.write_bytes(content)
    return {"url": f"/api/uploads/cheques/{fname}", "filename": fname}


@api_router.get("/uploads/cheques/{filename}")
async def serve_cheque(filename: str, user=Depends(current_user)):
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise HTTPException(404, "Not found")
    return FileResponse(str(path))


@api_router.get("/follow-ups")
async def list_follow_ups(user=Depends(current_user)):
    q: dict = {}
    if user["role"] == "field_officer":
        q["officer_id"] = user["id"]
    docs = await db.follow_ups.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [clean(d) for d in docs]


@api_router.post("/assignments")
async def assign(data: AssignmentInput, user=Depends(role_guard("owner", "head_officer"))):
    officer = await db.users.find_one({"id": data.officer_id, "role": "field_officer"}, {"_id": 0})
    if not officer:
        raise HTTPException(404, "Officer not found")
    r = await db.snapshot_parties.update_many(
        {"snapshot_id": data.snapshot_id, "id": {"$in": data.party_ids}},
        {"$set": {"assigned_officer_id": data.officer_id, "assigned_officer_name": officer["name"]}},
    )
    await db.audit_logs.insert_one({
        "id": secrets.token_hex(8), "action": "parties_assigned",
        "user_id": user["id"], "user_name": user["name"],
        "officer_id": data.officer_id, "officer_name": officer["name"],
        "count": len(data.party_ids), "created_at": now(),
    })
    return {"ok": True, "assigned": r.modified_count, "officer": officer["name"]}


@api_router.get("/dashboard")
async def dashboard(user=Depends(current_user)):
    snap = await db.snapshots.find_one({"active": True}, {"_id": 0})
    if not snap:
        return {
            "user": user, "snapshot": None,
            "metrics": {"total_outstanding": 0, "party_count": 0, "assigned_count": 0,
                        "unassigned_count": 0, "follow_up_count": 0, "due_today_count": 0,
                        "broken_promise_count": 0, "collected_total": 0},
            "outcomes": {"paid_full": 0, "paid_partial": 0, "not_paid": 0},
            "top_parties": [], "recent_follow_ups": [], "due_today": [], "broken_promises": [],
            "officer_performance": [],
        }
    parties = await db.snapshot_parties.find({"snapshot_id": snap["id"]}, {"_id": 0, "bills": 0}).to_list(3000)
    all_parties = parties
    if user["role"] == "field_officer":
        parties = [p for p in parties if p.get("assigned_officer_id") == user["id"]]

    total_outstanding = sum(p["total_outstanding"] for p in parties)
    assigned_count = len([p for p in parties if p.get("assigned_officer_id")])
    unassigned_count = len(parties) - assigned_count
    top_parties = sorted(parties, key=lambda x: x["total_outstanding"], reverse=True)[:8]

    fq: dict = {}
    if user["role"] == "field_officer":
        fq["officer_id"] = user["id"]
    follow_ups = await db.follow_ups.find(fq, {"_id": 0}).sort("created_at", -1).to_list(500)

    today = date.today().isoformat()
    due_today = [f for f in follow_ups if f.get("next_followup_date") == today]
    broken: list = []
    seen = set()
    for f in follow_ups:
        pid = f.get("party_id")
        if f.get("promise_date") and f["promise_date"] < today and pid not in seen:
            paid_after = any(x for x in follow_ups if x.get("party_id") == pid
                             and x.get("outcome") == "paid_full"
                             and x["created_at"] > f["created_at"])
            if not paid_after:
                broken.append(f)
                seen.add(pid)

    outcomes = {"paid_full": 0, "paid_partial": 0, "not_paid": 0}
    for f in follow_ups:
        outcomes[f.get("outcome", "not_paid")] = outcomes.get(f.get("outcome", "not_paid"), 0) + 1

    officer_perf = []
    if user["role"] in ("owner", "head_officer"):
        officers = await db.users.find({"role": "field_officer", "active": True}, {"_id": 0, "password_hash": 0}).to_list(50)
        for o in officers:
            fups = [f for f in follow_ups if f.get("officer_id") == o["id"]]
            assigned = len([p for p in all_parties if p.get("assigned_officer_id") == o["id"]])
            collected = sum(money(f.get("amount_received")) for f in fups
                            if f.get("outcome") in ("paid_full", "paid_partial"))
            officer_perf.append({
                "id": o["id"], "name": o["name"],
                "assigned_parties": assigned,
                "follow_ups": len(fups),
                "collected": round(collected, 2),
            })

    return {
        "user": user,
        "snapshot": clean(snap),
        "metrics": {
            "total_outstanding": round(total_outstanding, 2),
            "party_count": len(parties),
            "assigned_count": assigned_count,
            "unassigned_count": unassigned_count,
            "follow_up_count": len(follow_ups),
            "due_today_count": len(due_today),
            "broken_promise_count": len(broken),
            "collected_total": round(sum(money(f.get("amount_received")) for f in follow_ups
                                         if f.get("outcome") in ("paid_full", "paid_partial")), 2),
        },
        "outcomes": outcomes,
        "top_parties": [clean(p) for p in top_parties],
        "recent_follow_ups": [clean(f) for f in follow_ups[:10]],
        "due_today": [clean(f) for f in due_today[:15]],
        "broken_promises": [clean(f) for f in broken[:15]],
        "officer_performance": officer_perf,
    }


@api_router.get("/reports")
async def reports(user=Depends(role_guard("owner", "head_officer"))):
    snaps = await db.snapshots.find({}, {"_id": 0}).sort("uploaded_at", -1).to_list(50)
    fups = await db.follow_ups.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    audit = await db.audit_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"snapshots": [clean(s) for s in snaps],
            "follow_ups": [clean(f) for f in fups],
            "audit": [clean(a) for a in audit]}


async def seed():
    await db.users.create_index("email", unique=True)
    demos = [
        ("owner@sripati.local", "Owner", "owner"),
        ("head@sripati.local", "Meera Shah", "head_officer"),
        ("field1@sripati.local", "Rakesh Patel", "field_officer"),
        ("field2@sripati.local", "Neha Joshi", "field_officer"),
        ("field3@sripati.local", "Vijay Solanki", "field_officer"),
    ]
    for email, name, role in demos:
        if not await db.users.find_one({"email": email}):
            await db.users.insert_one({
                "id": secrets.token_hex(12), "email": email, "name": name, "role": role,
                "password_hash": hash_password("Sripati@123"),
                "active": True, "demo": True, "force_password_change": True,
                "created_at": now(),
            })


@app.on_event("startup")
async def startup():
    await seed()


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get("FRONTEND_URL", "https://mill-invoice-tracker.preview.emergentagent.com")],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown():
    client.close()
